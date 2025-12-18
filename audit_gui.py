import sys
import os
import time
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QPushButton, QLabel, QProgressBar, QMessageBox)
from PyQt6.QtCore import QThread, pyqtSignal, Qt

# --- LOGIC IMPORTS ---
from playwright.sync_api import sync_playwright
from read_data import get_links_from_excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# --- CONFIGURATION ---
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(base_path, "_sys_core")

# --- WORKER THREAD (The Engine) ---
class AuditWorker(QThread):
    status_update = pyqtSignal(str)
    progress_update = pyqtSignal(int)
    finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    validation_failed = pyqtSignal(str) 

    def __init__(self):
        super().__init__()
        self.start_permission = False 
        self.is_running = True

    def set_start_permission(self):
        self.start_permission = True

    def stop(self):
        self.is_running = False

    def run(self):
        try:
            self.status_update.emit("Loading Excel Data...")
            
            # 1. READ DATA
            file_kinnex = os.path.join(base_path, "Kinnex Revision Source.xlsx")
            file_quattro = os.path.join(base_path, "Quattro Revision Source.xlsx")
            
            links_kinnex = get_links_from_excel(file_kinnex, "Kinnex")
            links_quattro = get_links_from_excel(file_quattro, "Quattro")
            
            total_items = len(links_kinnex) + len(links_quattro)
            if total_items == 0:
                self.error_occurred.emit("No data found in Source files!")
                return

            # 2. LAUNCH BROWSER
            self.status_update.emit("Launching Browser...")
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False, args=["--start-maximized"])
                context = browser.new_context(no_viewport=True)
                page = context.new_page()

                # Go to login page
                first_url = links_kinnex[0]['url'] if links_kinnex else links_quattro[0]['url']
                try:
                    page.goto(first_url)
                except:
                    pass

                # --- THE RETRY LOOP ---
                while self.is_running:
                    
                    # A. WAIT FOR USER START
                    self.status_update.emit("WAITING: Log in to Laserfiche, then click START.")
                    while not self.start_permission:
                        if not self.is_running: break # Handle Abort
                        time.sleep(0.5)

                    if not self.is_running: break # Exit if Abort was clicked

                    # B. VALIDATE LOGIN (The Bouncer)
                    try:
                        page_title = page.title().lower()
                        current_url = page.url.lower()

                        if "login" in page_title or "sign in" in page_title or "login" in current_url:
                            # FAIL: Don't close browser. Just reset the GUI and wait again.
                            self.validation_failed.emit("Login Screen Detected!\n\nPlease log in to Laserfiche, then click 'START AUDIT' again.")
                            self.start_permission = False # Reset permission flag
                            continue # Jump back to step A (Wait loop)
                    except:
                        pass 

                    # If we passed validation, BREAK the retry loop and start scanning
                    break 
                # ----------------------

                # CHECK IF ABORTED BEFORE SCAN
                if not self.is_running:
                    browser.close()
                    self.finished.emit("User Aborted Audit.")
                    return

                # 3. START SCANNING
                self.status_update.emit("Audit Started...")
                processed_count = 0
                
                # Scan Kinnex
                kinnex_broken = []
                if self.is_running and links_kinnex:
                    kinnex_broken = self.scan_list(page, links_kinnex, "Kinnex", processed_count, total_items)
                    processed_count += len(links_kinnex)

                # Scan Quattro
                quattro_broken = []
                if self.is_running and links_quattro:
                    quattro_broken = self.scan_list(page, links_quattro, "Quattro", processed_count, total_items)

                browser.close()

                # CHECK IF ABORTED DURING SCAN
                if not self.is_running:
                    self.finished.emit("User Aborted Audit.")
                    return

                # 4. GENERATE REPORTS
                self.status_update.emit("Generating Reports...")
                
                if links_kinnex:
                    self.save_report(file_kinnex, "Kinnex", links_kinnex, kinnex_broken)
                if links_quattro:
                    self.save_report(file_quattro, "Quattro", links_quattro, quattro_broken)

                self.finished.emit("Audit Complete! Reports Saved.")

        except Exception as e:
            self.error_occurred.emit(str(e))

    def scan_list(self, page, links, name, current_count, total):
        broken = []
        for i, item in enumerate(links):
            if not self.is_running: break
            
            self.status_update.emit(f"Checking {name}: {item['text']}")
            progress_pct = int(((current_count + i) / total) * 100)
            self.progress_update.emit(progress_pct)

            is_dead = False
            try:
                page.goto(item['url'])
                page.wait_for_load_state("domcontentloaded")
                title = page.title()
                
                # REINFORCED CHECK: If we hit a login page mid-scan, count it as "Dead"
                title_lower = title.lower()
                if "entry not found" in title_lower or "404" in title_lower or "login" in title_lower:
                    is_dead = True
            except:
                is_dead = True
            
            if is_dead:
                broken.append(item['cell'])
        
        return broken

    def save_report(self, filename, name, all_links, broken_cells):
        if not os.path.exists(filename): return
        
        wb = load_workbook(filename, data_only=False)
        ws = wb.active
        
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        black_font = Font(color="000000")
        
        # Cleanup
        for item in all_links:
            cell = ws[item['cell']]
            cell.hyperlink = None
            cell.font = black_font
            
        # Highlight Broken
        for addr in broken_cells:
            rev_cell = ws[addr].offset(column=1)
            rev_cell.value = ""
            rev_cell.fill = yellow

        # Stamp Date
        now = datetime.now()
        if ws["K34"]: ws["K34"].value = now.strftime("%m/%d/%Y") 
        if ws["K35"]: ws["K35"].value = now.strftime("%I:%M %p")
        
        output_path = os.path.join(base_path, f"{name} Revisions.xlsx")
        wb.save(output_path)

# --- GUI WINDOW ---
class AuditDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Revision Auditor")
        self.resize(400, 300) 
        self.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout()
        layout.setSpacing(15)
        central.setLayout(layout)

        self.lbl_title = QLabel("Daily Revision Auditor")
        self.lbl_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #333;")
        self.lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.lbl_title)

        self.lbl_status = QLabel("Initializing...")
        self.lbl_status.setStyleSheet("font-size: 14px; color: #555;")
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_status.setWordWrap(True)
        layout.addWidget(self.lbl_status)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setStyleSheet("height: 25px;")
        layout.addWidget(self.progress)

        self.btn_start = QPushButton("START AUDIT")
        self.btn_start.setStyleSheet("""
            QPushButton {
                background-color: #2e7d32; 
                color: white; 
                font-size: 16px; 
                font-weight: bold;
                padding: 15px;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #1b5e20; }
            QPushButton:disabled { background-color: #ccc; }
        """)
        self.btn_start.setEnabled(False) 
        self.btn_start.clicked.connect(self.on_start_click)
        layout.addWidget(self.btn_start)

        self.btn_stop = QPushButton("ABORT")
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #757575; 
                color: white; 
                font-size: 14px; 
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #424242; }
            QPushButton:disabled { background-color: #e0e0e0; color: #9e9e9e; }
        """)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.on_stop_click)
        layout.addWidget(self.btn_stop)

        self.worker = AuditWorker()
        self.worker.status_update.connect(self.lbl_status.setText)
        self.worker.progress_update.connect(self.progress.setValue)
        self.worker.finished.connect(self.on_finished)
        self.worker.error_occurred.connect(self.on_error)
        self.worker.validation_failed.connect(self.on_validation_failed) 
        
        self.worker.start()
        
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(True) 
        
        self.center_and_offset()

    def center_and_offset(self):
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x + 450, y) 

    def on_start_click(self):
        self.btn_start.setText("SCANNING...")
        self.btn_start.setEnabled(False) 
        self.btn_start.setStyleSheet("background-color: #d32f2f; color: white; padding: 15px; font-weight: bold;")
        self.worker.set_start_permission()

    def on_stop_click(self):
        self.lbl_status.setText("Aborting...")
        self.btn_stop.setEnabled(False)
        self.btn_start.setEnabled(False)
        self.worker.stop()

    def on_validation_failed(self, msg):
        QMessageBox.warning(self, "Login Required", msg)
        
        self.lbl_status.setText("Waiting for Login...")
        self.btn_start.setText("START AUDIT")
        self.btn_start.setEnabled(True)
        self.btn_start.setStyleSheet("""
            QPushButton {
                background-color: #2e7d32; 
                color: white; 
                font-size: 16px; 
                font-weight: bold;
                padding: 15px;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #1b5e20; }
        """)

    def on_finished(self, msg):
        self.lbl_status.setText(msg)
        self.progress.setValue(100) if "Complete" in msg else self.progress.setValue(0)
        
        self.btn_start.setText("CLOSE")
        self.btn_start.setStyleSheet("background-color: #1976d2; color: white; padding: 15px;")
        self.btn_start.setEnabled(True)
        
        try: self.btn_start.clicked.disconnect() 
        except: pass
        self.btn_start.clicked.connect(self.close)
        
        self.btn_stop.setEnabled(False)

    def on_error(self, msg):
        self.lbl_status.setText("Error!")
        QMessageBox.critical(self, "Error", msg)
        self.on_finished("Error Occurred.")

    def closeEvent(self, event):
        self.worker.stop()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AuditDashboard()
    window.show()
    sys.exit(app.exec())