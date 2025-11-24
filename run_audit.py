import os
import sys

# --- FIX FOR PYINSTALLER + PLAYWRIGHT ---
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(base_path, "browsers")
# ----------------------------------------

from playwright.sync_api import sync_playwright
from read_data import get_links_from_excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# --- CORE FUNCTIONS ---

def scan_links(page, links, customer_name):
    if not links:
        print(f"Skipping {customer_name} (No links found).")
        return []

    print(f"\n--- SCANNING {customer_name} ({len(links)} docs) ---")
    broken_cells = []

    for item in links:
        print(f"Checking {item['text']}...", end="", flush=True)
        
        is_dead = False
        try:
            page.goto(item['url'])
            page.wait_for_load_state("domcontentloaded")
            title = page.title()
            
            if "Entry not found" in title or "Application Error" in title or "404" in title:
                    is_dead = True
                    print(f" -> DEAD LINK (Highlighting)")
            else:
                print(f" -> OK")
            
        except Exception as e:
            is_dead = True
            print(f" -> Failed to load: {e}")
        
        if is_dead:
            broken_cells.append(item['cell'])
            
    return broken_cells

def generate_report(source_filename, customer_name, all_links, broken_cells):
    print(f"Generating Report for {customer_name}...")
    
    if not os.path.exists(source_filename):
        print(f"Error: Source file {source_filename} not found.")
        return

    wb = load_workbook(source_filename, data_only=False)
    ws = wb.active
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")
    
    for item in all_links:
        cell = ws[item['cell']]
        cell.hyperlink = None
        cell.font = black_font
    
    count = 0
    for cell_addr in broken_cells:
        link_cell = ws[cell_addr]
        rev_cell = link_cell.offset(column=1)
        rev_cell.value = ""
        rev_cell.fill = yellow_fill
        count += 1
        
    now = datetime.now()
    if ws["K34"]: ws["K34"].value = now.strftime("%m/%d/%Y") 
    if ws["K35"]: ws["K35"].value = now.strftime("%I:%M %p")

    output_filename = f"{customer_name} Revisions.xlsx"
    wb.save(output_filename)
    print(f" -> Saved: {output_filename} ({count} broken links cleared & highlighted)")

# --- MAIN EXECUTION FLOW ---

def run_daily_audit():
    print("="*50)
    print("   DAILY REVISION AUDIT: KINNEX & QUATTRO")
    print("="*50)

    file_kinnex = "Kinnex Revision Source.xlsx"
    file_quattro = "Quattro Revision Source.xlsx"
    
    links_kinnex = get_links_from_excel(file_kinnex, "Kinnex")
    links_quattro = get_links_from_excel(file_quattro, "Quattro")
    
    if not links_kinnex and not links_quattro:
        print("No data found in either file. Exiting.")
        return

    kinnex_broken = []
    quattro_broken = []

    with sync_playwright() as p:
        print("\nLaunching Browser...")
        # Start maximized
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        first_url = links_kinnex[0]['url'] if links_kinnex else links_quattro[0]['url']
        
        print(f"\nOPENING LOGIN PAGE...")
        try:
            page.goto(first_url)
        except:
            pass

        # --- THE RELIABLE PAUSE ---
        print("\n" + "#"*60)
        print("ACTION REQUIRED:")
        print("1. The Browser is now open.")
        print("2. Please LOG IN to Laserfiche manually.")
        print("3. Once you see the document, CLICK INSIDE THIS BLACK WINDOW.")
        print("4. Press ENTER on your keyboard to start the audit.")
        print("#"*60)
        
        input("\n>>> PRESS ENTER HERE TO START SCANS <<<")
        print("Starting batch processing...")

        # --- RUN SCANS ---
        kinnex_broken = scan_links(page, links_kinnex, "Kinnex")
        quattro_broken = scan_links(page, links_quattro, "Quattro")

        print("\nClosing Browser...")
        browser.close()

    print("\n" + "-"*30)
    if links_kinnex:
        generate_report(file_kinnex, "Kinnex", links_kinnex, kinnex_broken)
    if links_quattro:
        generate_report(file_quattro, "Quattro", links_quattro, quattro_broken)
    print("-"*30)
    print("ALL AUDITS COMPLETE.")

if __name__ == "__main__":
    run_daily_audit()